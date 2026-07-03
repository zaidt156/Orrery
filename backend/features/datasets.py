"""Imported data sources, BI-style: uploaded CSV/Excel files and REST APIs (JSON) become real tables
in the orrery_datasets schema of Orrery's own Postgres, exposed to dashboards through one built-in
"Workspace datasets" connection that is scoped to that schema — imports can never see app tables
(chats, projects) and app queries can't be widened by imports.

Safety: table/column identifiers are derived only from a sanitized [a-z0-9_] slug and quoted; all
values are inserted with bound parameters; API auth headers live in the OS keychain, never in the DB.
API fetches are user-initiated (a person typed the URL), size-capped, and parsed defensively."""
from __future__ import annotations

import base64
import io
import json
import logging
import re
import uuid

from sqlalchemy import select, text

from backend.core.database import get_engine, get_sessionmaker, resolve_database_url
from backend.core.models import DataConnection, Dataset
from backend.security import secrets

log = logging.getLogger("orrery.datasets")

SCHEMA = "orrery_datasets"
MAX_ROWS = 50_000
MAX_COLS = 100
MAX_API_BYTES = 8_000_000
_BATCH = 500

_IDENT = re.compile(r"[^a-z0-9_]+")


def _slug(name: str, prefix: str = "", limit: int = 60) -> str:
    s = _IDENT.sub("_", (name or "").strip().lower()).strip("_") or "col"
    if s[0].isdigit():
        s = f"c_{s}"
    return (prefix + s)[:limit]


def _quote(ident: str) -> str:
    return '"' + ident.replace('"', '""') + '"'


def _column_names(header: list) -> list[str]:
    out, seen = [], set()
    for i, raw in enumerate(header[:MAX_COLS]):
        base = _slug(str(raw) or f"col_{i + 1}")
        name = base
        n = 2
        while name in seen:
            name = f"{base}_{n}"
            n += 1
        seen.add(name)
        out.append(name)
    return out


def _infer_types(rows: list[list], n_cols: int) -> list[str]:
    """DOUBLE PRECISION where every non-empty value parses as a number, else TEXT."""
    numeric = [True] * n_cols
    for row in rows[:2000]:
        for i in range(n_cols):
            v = row[i] if i < len(row) else None
            if v in (None, ""):
                continue
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                continue
            try:
                float(str(v).replace(",", ""))
            except (ValueError, TypeError):
                numeric[i] = False
    return ["DOUBLE PRECISION" if numeric[i] else "TEXT" for i in range(n_cols)]


def _cell(value, sql_type: str):
    if value in (None, ""):
        return None
    if sql_type == "DOUBLE PRECISION":
        try:
            return float(str(value).replace(",", ""))
        except (ValueError, TypeError):
            return None
    return str(value)


async def _materialize(table: str, header: list, rows: list[list], schema: str = SCHEMA) -> int:
    """(Re)create schema.table from header+rows. Identifiers sanitized, values parameterized."""
    cols = _column_names(header)
    types = _infer_types(rows, len(cols))
    qtable = f"{_quote(schema)}.{_quote(table)}"
    col_defs = ", ".join(f"{_quote(c)} {t}" for c, t in zip(cols, types))
    params_list = [
        {f"p{i}": _cell(row[i] if i < len(row) else None, types[i]) for i in range(len(cols))}
        for row in rows[:MAX_ROWS]
    ]
    placeholders = ", ".join(f":p{i}" for i in range(len(cols)))
    engine = get_engine()
    async with engine.begin() as conn:
        await conn.execute(text(f"DROP TABLE IF EXISTS {qtable}"))
        await conn.execute(text(f"CREATE TABLE {qtable} ({col_defs})"))
        insert = text(f"INSERT INTO {qtable} ({', '.join(_quote(c) for c in cols)}) VALUES ({placeholders})")
        for start in range(0, len(params_list), _BATCH):
            batch = params_list[start:start + _BATCH]
            if batch:
                await conn.execute(insert, batch)
    return len(params_list)


# --- dataset workspaces (each = its own schema + its own selectable connection) -------------------

def _attach_url(cid: str) -> None:
    if not secrets.get_secret(f"conn:{cid}"):
        url = resolve_database_url()
        if url:
            secrets.set_secret(f"conn:{cid}", url)


async def ensure_connection() -> str:
    """Find-or-create the DEFAULT workspace connection (schema orrery_datasets). Returns its id."""
    async with get_sessionmaker()() as s:
        row = (await s.execute(
            select(DataConnection).where(DataConnection.kind == "datasets")
            .order_by(DataConnection.created_at)
        )).scalars().first()
        if row is not None:
            cid = str(row.id)
        else:
            row = DataConnection(name="Workspace datasets", display="imported files & APIs",
                                 kind="datasets", db_schema=SCHEMA)
            s.add(row)
            await s.commit()
            await s.refresh(row)
            cid = str(row.id)
    _attach_url(cid)
    return cid


async def create_workspace(name: str) -> dict:
    """A new dataset workspace: its own schema + its own connection, selectable in dashboards."""
    schema = ("orrery_ws_" + _slug(name, limit=40)).lower()
    engine = get_engine()
    async with engine.begin() as conn:
        await conn.execute(text(f"CREATE SCHEMA IF NOT EXISTS {_quote(schema)}"))
    async with get_sessionmaker()() as s:
        existing = (await s.execute(select(DataConnection).where(DataConnection.db_schema == schema))).scalars().first()
        if existing is not None:
            return {"id": str(existing.id), "name": existing.name, "schema": schema}
        row = DataConnection(name=(name.strip() or "workspace")[:120], display=f"workspace · {name.strip()[:60]}",
                             kind="datasets", db_schema=schema)
        s.add(row)
        await s.commit()
        await s.refresh(row)
        cid = str(row.id)
    _attach_url(cid)
    return {"id": cid, "name": name.strip() or "workspace", "schema": schema}


async def list_workspaces() -> list[dict]:
    await ensure_connection()  # the default workspace always exists
    async with get_sessionmaker()() as s:
        rows = (await s.execute(
            select(DataConnection).where(DataConnection.kind == "datasets").order_by(DataConnection.created_at)
        )).scalars().all()
        return [{"id": str(r.id), "name": r.name, "schema": r.db_schema or SCHEMA} for r in rows]


async def _workspace_schema(workspace_id: str | None) -> str:
    if not workspace_id:
        await ensure_connection()
        return SCHEMA
    async with get_sessionmaker()() as s:
        row = await s.get(DataConnection, uuid.UUID(workspace_id))
        if row is None or row.kind != "datasets":
            raise ValueError("Workspace not found.")
        return row.db_schema or SCHEMA


# --- parsing --------------------------------------------------------------------------------------

def _parse_csv(text_content: str) -> tuple[list, list[list]]:
    import csv as _csv
    rows = list(_csv.reader(io.StringIO(text_content)))
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if len(rows) < 2:
        raise ValueError("The file needs a header row and at least one data row.")
    return rows[0], rows[1:MAX_ROWS + 1]


def _parse_xlsx(data: bytes) -> tuple[list, list[list]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(v not in (None, "") for v in row):
            rows.append(list(row))
        if len(rows) > MAX_ROWS:
            break
    if len(rows) < 2:
        raise ValueError("The sheet needs a header row and at least one data row.")
    return rows[0], rows[1:]


def _flatten_json(payload) -> tuple[list, list[list]]:
    """Find the main list of records in a JSON payload; nested values become JSON strings."""
    records = None
    if isinstance(payload, list):
        records = payload
    elif isinstance(payload, dict):
        for key in ("data", "items", "results", "records", "rows"):
            if isinstance(payload.get(key), list):
                records = payload[key]
                break
        if records is None:  # fall back to the first list value anywhere at the top level
            records = next((v for v in payload.values() if isinstance(v, list)), None)
        if records is None:
            records = [payload]  # single object → one-row table
    if not records:
        raise ValueError("The API response contained no list of records.")
    records = [r for r in records if isinstance(r, dict)][:MAX_ROWS]
    if not records:
        raise ValueError("The API records aren't JSON objects.")
    header: list[str] = []
    for r in records[:200]:
        for k in r.keys():
            if k not in header and len(header) < MAX_COLS:
                header.append(k)
    rows = [[(json.dumps(v) if isinstance(v, (dict, list)) else v) for v in (r.get(k) for k in header)]
            for r in records]
    return header, rows


async def _fetch_api(url: str, headers: dict[str, str] | None) -> tuple[list, list[list]]:
    import httpx

    from backend.features import team
    from backend.security import netguard
    # SSRF guard: metadata/link-local always blocked; in team mode members can't probe the host's
    # LAN or loopback through imports (solo users may legitimately import from their own local APIs).
    try:
        url = netguard.validate_fetch_url(url, allow_private=not await team.team_mode())
    except netguard.UnsafeUrlError as exc:
        raise ValueError(str(exc))
    async with httpx.AsyncClient(timeout=25, follow_redirects=True) as client:
        resp = await client.get(url, headers=headers or {})
        resp.raise_for_status()
        if len(resp.content) > MAX_API_BYTES:
            raise ValueError("The API response is too large (8 MB cap).")
        try:
            payload = resp.json()
        except ValueError:
            raise ValueError("The API didn't return JSON.")
    return _flatten_json(payload)


# --- registry CRUD ---------------------------------------------------------------------------------

def _headers_secret(did: str) -> str:
    return f"dataset_headers:{did}"


def _dict(d: Dataset) -> dict:
    return {
        "id": str(d.id), "name": d.name, "table": d.table_name, "kind": d.kind,
        "schema": getattr(d, "db_schema", SCHEMA) or SCHEMA,
        "source": d.source or "", "rows": int(d.row_count or 0),
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


async def list_datasets() -> list[dict]:
    async with get_sessionmaker()() as s:
        rows = (await s.execute(select(Dataset).order_by(Dataset.created_at))).scalars().all()
        return [_dict(r) for r in rows]


async def _register(name: str, kind: str, source: str, header: list, rows: list[list],
                    workspace_id: str | None = None) -> dict:
    if len(rows) > MAX_ROWS:
        rows = rows[:MAX_ROWS]
    schema = await _workspace_schema(workspace_id)
    base = _slug(name, prefix="ds_")
    table = base
    async with get_sessionmaker()() as s:
        n = 2
        while (await s.execute(select(Dataset).where(Dataset.table_name == table))).scalars().first():
            table = f"{base}_{n}"
            n += 1
    count = await _materialize(table, header, rows, schema=schema)
    async with get_sessionmaker()() as s:
        row = Dataset(name=(name.strip() or "dataset")[:120], table_name=table, kind=kind,
                      db_schema=schema, source=(source or "")[:500] or None, row_count=count)
        s.add(row)
        await s.commit()
        await s.refresh(row)
        return _dict(row)


_GSHEET = re.compile(r"docs\.google\.com/spreadsheets/d/([A-Za-z0-9_-]+)")


async def create_from_file(name: str, filename: str, content: str, workspace_id: str | None = None) -> dict:
    """content: raw text for CSV/JSON, base64 for Excel (the UI sends the right one by extension)."""
    fname = (filename or "").lower()
    if fname.endswith((".xlsx", ".xls", ".xlsm")):
        header, rows = _parse_xlsx(base64.b64decode(content))
    elif fname.endswith(".json"):
        header, rows = _flatten_json(json.loads(content))
    else:
        header, rows = _parse_csv(content)
    return await _register(name or filename, "file", filename, header, rows, workspace_id)


async def create_from_api(name: str, url: str, headers: dict[str, str] | None = None,
                          workspace_id: str | None = None) -> dict:
    # A shared Google Sheets link imports via its CSV export (no auth; sheet must be link-visible).
    gs = _GSHEET.search(url or "")
    if gs:
        import httpx
        export = f"https://docs.google.com/spreadsheets/d/{gs.group(1)}/export?format=csv"
        async with httpx.AsyncClient(timeout=25, follow_redirects=True) as client:
            resp = await client.get(export)
            resp.raise_for_status()
            if len(resp.content) > MAX_API_BYTES:
                raise ValueError("The sheet is too large (8 MB cap).")
        header, rows = _parse_csv(resp.text)
        return await _register(name or "google sheet", "api", export, header, rows, workspace_id)
    header, rows = await _fetch_api(url, headers)
    out = await _register(name or url, "api", url, header, rows, workspace_id)
    if headers:
        secrets.set_secret(_headers_secret(out["id"]), json.dumps(headers))
    return out


async def refresh_dataset(did: str) -> dict:
    async with get_sessionmaker()() as s:
        row = await s.get(Dataset, uuid.UUID(did))
        if row is None:
            raise ValueError("Dataset not found.")
        if row.kind != "api" or not row.source:
            raise ValueError("Only API datasets can be refreshed — re-upload the file instead.")
        table, url, schema = row.table_name, row.source, (row.db_schema or SCHEMA)
    raw = secrets.get_secret(_headers_secret(did))
    headers = json.loads(raw) if raw else None
    if "docs.google.com/spreadsheets" in url:
        import httpx
        async with httpx.AsyncClient(timeout=25, follow_redirects=True) as client:
            resp = await client.get(url)
            resp.raise_for_status()
        header, rows = _parse_csv(resp.text)
    else:
        header, rows = await _fetch_api(url, headers)
    count = await _materialize(table, header, rows, schema=schema)
    async with get_sessionmaker()() as s:
        row = await s.get(Dataset, uuid.UUID(did))
        if row is not None:
            row.row_count = count
            await s.commit()
            await s.refresh(row)
            return _dict(row)
    raise ValueError("Dataset not found.")


async def delete_dataset(did: str) -> bool:
    async with get_sessionmaker()() as s:
        row = await s.get(Dataset, uuid.UUID(did))
        if row is None:
            return False
        table, schema = row.table_name, (row.db_schema or SCHEMA)
        await s.delete(row)
        await s.commit()
    engine = get_engine()
    async with engine.begin() as conn:
        await conn.execute(text(f"DROP TABLE IF EXISTS {_quote(schema)}.{_quote(table)}"))
    secrets.delete_secret(_headers_secret(did))
    return True
